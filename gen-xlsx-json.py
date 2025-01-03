import json
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def create_excel_from_json(json_file, excel_file):
    """
    Creates an Excel file (.xlsx) from a JSON data file with specified formatting, including parts.

    Args:
        json_file: Path to the input JSON file (data.json).
        excel_file: Path to the output Excel file (data.xlsx).
    """
    # Type abbreviation mapping
    type_map = {
        "determiner": "det",
        "verb": "v",
        "auxiliary verb": "aux",
        "noun": "n",
        "conjunction": "conj",
        "preposition": "prep",
        "adverb": "adv",
        "adjective": "adj",
        "pronoun": "pron"
    }

    # Create a new Excel workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Words"

    # Write the header row
    header = ["Word", "Types", "Parts", "IPA", "Vietnamese"]
    ws.append(header)

    # Apply font size 12, bold and center alignment to the header row
    for cell in ws[1]:  # ws[1] represents the first row (header)
        cell.font = Font(size=12, bold=True)

    # Load JSON data
    with open(json_file, 'r', encoding='utf-8') as f_json:
        data = json.load(f_json)

        # Fill in the data rows
        for item in data:
            word = item["word"]
            ipa = item["ipa"]
            parts = item.get("parts", [])  # Get parts or default to empty list
            parts_str = "; ".join(parts)  # convert parts list to parts string

            # Combine meanings from different types
            all_vietnamese_meanings = []
            types = []
            for pos, meanings in item["meaning"].items():
                types.append(type_map.get(pos, pos))
                all_vietnamese_meanings.extend(meanings)

            # Create the type string (e.g., "n, v, adj")
            type_str = ", ".join(types)
            # Create the Vietnamese meanings string (e.g., "cái; con; người")
            vietnamese_str = "; ".join(all_vietnamese_meanings)

            # Write the data row, include the parts string
            ws.append([word, type_str, parts_str, ipa, vietnamese_str])

    # Adjust column width to fit content for all columns
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column:
            if cell.value:  # Check if the cell has a value
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        # Set column width based on content length, without unnecessary padding
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length  # No padding

    # Save the Excel file
    wb.save(excel_file)

# Example usage
json_file = "data.json"
excel_file = "data.xlsx"

create_excel_from_json(json_file, excel_file)
